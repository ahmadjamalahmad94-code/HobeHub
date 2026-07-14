# صفحة «الحضور والغياب» — من حضر كل أسبوع:
#   • البطاقات: من سجل البطاقات المحلّيّ (أيّ يوم/ساعة أخذ بطاقة).
#   • الإنترنت: جلسات الريديوس (متى دخل/خرج كل يوم ومدّته).
# مع اسم/جوّال/نوع/تخصص/جامعة-شركة، فلترة بمدى زمنيّ، وتصدير CSV.

import csv as _csv
import io as _io
import re as _re
from datetime import datetime as _dt
from flask import render_template, request, Response


def _norm_phone(v):
    """آخر ٩ أرقام من رقم الجوّال (يتجاوز اختلاف الصيفر/المقدّمة الدوليّة)."""
    digits = _re.sub(r"\D", "", str(v or ""))
    return digits[-9:] if len(digits) >= 9 else digits

_TYPE_LABEL = {"university": "جامعي", "freelancer": "عمل حر", "tawjihi": "توجيهي"}
_AR_DAYS = ["الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]


def _spec_of(b):
    t = (b.get("user_type") or "").lower()
    if t == "university":
        return b.get("university_specialization") or ""
    if t == "freelancer":
        return b.get("freelancer_specialization") or ""
    if t == "tawjihi":
        return b.get("tawjihi_branch") or ""
    return ""


def _org_of(b):
    t = (b.get("user_type") or "").lower()
    if t == "university":
        return b.get("university_name") or ""
    if t == "freelancer":
        return b.get("freelancer_company") or ""
    if t == "tawjihi":
        return str(b.get("tawjihi_year") or "")
    return ""


def _day_name(date_str):
    try:
        return _AR_DAYS[_dt.fromisoformat(str(date_str)[:10]).weekday()]
    except Exception:
        return ""


def _hm(value):
    """يستخرج HH:MM من طابع زمنيّ نصّيّ."""
    s = str(value or "")
    if "T" in s:
        s = s.split("T", 1)[1]
    elif " " in s:
        s = s.split(" ", 1)[1]
    return s[:5] if len(s) >= 5 else s


def _attendance_range():
    d_from = clean_csv_value(request.args.get("from"))
    d_to = clean_csv_value(request.args.get("to"))
    today = today_local()
    if not d_from:
        d_from = str(get_week_start(today))
    if not d_to:
        d_to = str(today)
    return d_from, d_to


def _cards_attendance(d_from, d_to):
    rows = query_all(
        """
        SELECT l.usage_date, l.usage_time, l.card_type, l.usage_reason,
               b.full_name, b.phone, b.user_type,
               b.university_name, b.university_specialization,
               b.freelancer_company, b.freelancer_specialization,
               b.tawjihi_year, b.tawjihi_branch
        FROM beneficiary_usage_logs l
        JOIN beneficiaries b ON b.id = l.beneficiary_id
        WHERE l.usage_date >= %s AND l.usage_date <= %s
        ORDER BY l.usage_time DESC
        LIMIT 3000
        """,
        [d_from, d_to],
    )
    out = []
    for r in rows:
        out.append({
            "full_name": r.get("full_name") or "—",
            "phone": r.get("phone") or "—",
            "type_label": _TYPE_LABEL.get((r.get("user_type") or "").lower(), r.get("user_type") or "—"),
            "spec": _spec_of(r) or "—",
            "org": _org_of(r) or "—",
            "day_name": _day_name(r.get("usage_date")),
            "date": str(r.get("usage_date") or "")[:10],
            "time": _hm(r.get("usage_time")),
            "card_type": r.get("card_type") or "—",
            "reason": r.get("usage_reason") or "—",
        })
    return out


def _fmt_dur(secs):
    try:
        secs = int(secs or 0)
    except (TypeError, ValueError):
        secs = 0
    return ("%dس %dد" % (secs // 3600, (secs % 3600) // 60)) if secs else "—"


def _internet_attendance(d_from, d_to):
    from app.services.radius_client import get_radius_client, is_api_under_development
    if is_api_under_development():
        return []
    try:
        client = get_radius_client()
    except Exception:
        return []

    # خريطة اسم المستخدم → مستفيد مسجّل (بالاسم الخارجيّ أو الجوّال المُطبّع).
    ben_map = {}
    try:
        for b in query_all(
            "SELECT b.id, b.full_name, b.phone, b.user_type, b.university_name, "
            "b.university_specialization, b.freelancer_company, b.freelancer_specialization, "
            "b.tawjihi_year, b.tawjihi_branch, r.external_username "
            "FROM beneficiaries b "
            "LEFT JOIN beneficiary_radius_accounts r ON r.beneficiary_id=b.id"):
            if b.get("external_username"):
                ben_map["u:" + str(b["external_username"]).strip().lower()] = b
            _np = _norm_phone(b.get("phone"))
            if _np:
                ben_map.setdefault("p:" + _np, b)
    except Exception:
        ben_map = {}

    def _lookup(uname):
        b = ben_map.get("u:" + uname.lower())
        if not b:
            np = _norm_phone(uname)
            b = ben_map.get("p:" + np) if np else None
        return b

    def _row(b, uname, day, start, stop, secs, online=False):
        return {
            "full_name": b.get("full_name") or "—",
            "phone": b.get("phone") or uname or "—",
            "username": uname,
            "type_label": _TYPE_LABEL.get((b.get("user_type") or "").lower(), b.get("user_type") or "—"),
            "spec": _spec_of(b) or "—",
            "org": _org_of(b) or "—",
            "day_name": _day_name(day),
            "date": day,
            "start": start,
            "stop": stop,
            "duration": _fmt_dur(secs),
            "online": online,
        }

    # اجمع الجلسات المطبّقة (متصل الآن + مُغلَقة) ثم جمّعها: صفّ لكل (شخص، يوم).
    flat, seen = [], set()
    today_str = str(today_local())

    if d_from <= today_str <= d_to:
        try:
            online = client.get_online_users() or []
        except Exception:
            online = []
        for s in (online or []):
            if not isinstance(s, dict):
                continue
            uname = str(s.get("username") or "").strip()
            b = _lookup(uname) if uname else None
            if not b:
                continue
            sid = str(s.get("session_id") or s.get("acctsessionid") or "")
            if sid:
                seen.add(sid)
            try:
                secs = int(s.get("running_seconds") or s.get("session_time") or 0)
            except (TypeError, ValueError):
                secs = 0
            start_raw = s.get("started_at") or s.get("acctstarttime") or ""
            flat.append((_norm_phone(b.get("phone")) or uname, today_str, b, uname,
                         _hm(start_raw) if start_raw else "", "", True, secs))

    # ترقيم عبر الصفحات (الأحدث أوّلًا) حتى تغطية بداية المدى — لا اليوم فقط.
    sessions = []
    if hasattr(client, "get_accounting_sessions"):
        offset = 0
        for _pg in range(12):  # سقف ~6000 جلسة
            try:
                page = client.get_accounting_sessions(limit=500, offset=offset) or []
            except Exception:
                page = []
            if not page:
                break
            sessions.extend(page)
            oldest = ""
            for _s in page:
                if isinstance(_s, dict):
                    _dd = str(_s.get("started_at") or _s.get("acctstarttime") or "")[:10]
                    if _dd and (not oldest or _dd < oldest):
                        oldest = _dd
            if len(page) < 500 or (oldest and oldest < d_from):
                break
            offset += 500
    for s in (sessions or []):
        if not isinstance(s, dict):
            continue
        uname = str(s.get("username") or "").strip()
        b = _lookup(uname) if uname else None
        if not b:
            continue
        sid = str(s.get("session_id") or s.get("acctsessionid") or "")
        if sid and sid in seen:
            continue
        start = s.get("started_at") or s.get("acctstarttime") or ""
        stop = s.get("stopped_at") or s.get("acctstoptime") or ""
        day = str(start)[:10]
        if day and (day < d_from or day > d_to):
            continue
        try:
            secs = int(s.get("duration_sec") or s.get("acctsessiontime") or 0)
        except (TypeError, ValueError):
            secs = 0
        flat.append((_norm_phone(b.get("phone")) or uname, day, b, uname,
                     _hm(start), _hm(stop) if stop else "", False, secs))

    # تجميع: أوّل دخول ← آخر خروج + إجماليّ المدّة لكل (شخص، يوم).
    agg = {}
    for key, day, b, uname, start_hm, stop_hm, online, secs in flat:
        if not day:
            continue
        g = agg.get((key, day))
        if not g:
            g = {"b": b, "uname": uname, "first": "", "last": "", "online": False, "total": 0}
            agg[(key, day)] = g
        if start_hm and (not g["first"] or start_hm < g["first"]):
            g["first"] = start_hm
        if online:
            g["online"] = True
        elif stop_hm and stop_hm > g["last"]:
            g["last"] = stop_hm
        g["total"] += secs

    out = []
    for (key, day), g in agg.items():
        stop_disp = "متصل الآن" if g["online"] else (g["last"] or "—")
        out.append(_row(g["b"], g["uname"], day, g["first"] or "—", stop_disp, g["total"], online=g["online"]))

    out.sort(key=lambda x: (1 if x.get("online") else 0, x["date"], x["start"]), reverse=True)
    return out


def _csv_response(filename, headers, rows_iter):
    buf = _io.StringIO()
    buf.write("﻿")  # BOM كي تفتح Excel العربيّة صحيحًا
    w = _csv.writer(buf)
    w.writerow(headers)
    for row in rows_iter:
        w.writerow(row)
    return Response(
        buf.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _week_dates(d_from, d_to):
    """قائمة التواريخ من d_from إلى d_to (شاملة، بحدّ أقصى 40 يومًا)."""
    from datetime import timedelta as _td
    try:
        a = _dt.fromisoformat(str(d_from)[:10])
        b = _dt.fromisoformat(str(d_to)[:10])
    except Exception:
        return []
    if b < a:
        a, b = b, a
    days, cur = [], a
    while cur <= b and len(days) <= 40:
        days.append(cur.date().isoformat())
        cur += _td(days=1)
    return days


def _matrix_people(rows):
    """يجمع الصفوف حسب الشخص (بالجوّال المُطبّع) مع مجموعة أيّام حضوره."""
    people = {}
    for r in rows:
        key = _norm_phone(r.get("phone")) or (r.get("username") or r.get("full_name") or "?")
        p = people.setdefault(key, {
            "full_name": r.get("full_name") or "—", "phone": r.get("phone") or "—",
            "type_label": r.get("type_label") or "—", "spec": r.get("spec") or "—",
            "org": r.get("org") or "—", "days": set()})
        d = str(r.get("date") or "")[:10]
        if d:
            p["days"].add(d)
    return people


def _xlsx_matrix_response(filename, days, people):
    """ملفّ Excel: صفّ لكل شخص، عمود لكل يوم (✓ حاضر / فارغ)، وإجماليّ."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "الحضور"
    ws.sheet_view.rightToLeft = True
    day_labels = ["%s %s/%s" % (_day_name(d), d[8:10], d[5:7]) for d in days]
    header = ["الاسم", "الجوال", "النوع", "التخصص", "الجامعة/الشركة"] + day_labels + ["إجمالي الأيام"]
    ws.append(header)
    hfill = PatternFill("solid", fgColor="1E1E1E")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill, c.font, c.alignment, c.border = hfill, hfont, center, border
    for p in sorted(people.values(), key=lambda x: x["full_name"]):
        row = [p["full_name"], p["phone"], p["type_label"], p["spec"], p["org"]]
        cnt = 0
        for d in days:
            hit = d in p["days"]
            row.append("✓" if hit else "")
            cnt += 1 if hit else 0
        row.append(cnt)
        ws.append(row)
    widths = [26, 14, 10, 20, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(6, 6 + len(days) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.border = border
            if c.column >= 6:
                c.alignment = center
    ws.freeze_panes = "F2"
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/admin/attendance", methods=["GET"])
@login_required
@permission_required("view")
def admin_attendance_page():
    d_from, d_to = _attendance_range()
    export = clean_csv_value(request.args.get("export"))

    if export in ("cards_matrix", "internet_matrix"):
        days = _week_dates(d_from, d_to)
        src = _cards_attendance(d_from, d_to) if export == "cards_matrix" else _internet_attendance(d_from, d_to)
        people = _matrix_people(src)
        label = "cards" if export == "cards_matrix" else "internet"
        return _xlsx_matrix_response(f"attendance-{label}-matrix-{d_from}_{d_to}.xlsx", days, people)

    if export == "cards":
        rows = _cards_attendance(d_from, d_to)
        return _csv_response(
            f"attendance-cards-{d_from}_{d_to}.csv",
            ["الاسم", "الجوال", "النوع", "التخصص", "الجامعة/الشركة", "اليوم", "التاريخ", "الوقت", "نوع البطاقة", "السبب"],
            ([r["full_name"], r["phone"], r["type_label"], r["spec"], r["org"],
              r["day_name"], r["date"], r["time"], r["card_type"], r["reason"]] for r in rows),
        )
    if export == "internet":
        rows = _internet_attendance(d_from, d_to)
        return _csv_response(
            f"attendance-internet-{d_from}_{d_to}.csv",
            ["الاسم", "الجوال", "اسم المستخدم", "النوع", "التخصص", "الجامعة/الشركة", "اليوم", "التاريخ", "الدخول", "الخروج", "المدة"],
            ([r["full_name"], r["phone"], r["username"], r["type_label"], r["spec"], r["org"],
              r["day_name"], r["date"], r["start"], r["stop"], r["duration"]] for r in rows),
        )

    cards_rows = _cards_attendance(d_from, d_to)
    internet_rows = _internet_attendance(d_from, d_to)
    return render_template(
        "admin/attendance/list.html",
        d_from=d_from, d_to=d_to,
        cards_rows=cards_rows, internet_rows=internet_rows,
        cards_count=len(cards_rows), internet_count=len(internet_rows),
    )
